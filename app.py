from cs50 import SQL
from flask import Flask, flash, jsonify, redirect, render_template, request, session
from flask_session import Session
from tempfile import mkdtemp
from werkzeug.exceptions import default_exceptions, HTTPException, InternalServerError
from werkzeug.security import check_password_hash, generate_password_hash
from datetime import datetime
from functools import wraps

import pandas as pd
from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

#App configure
app = Flask(__name__)

#template auto reload
app.config["TEMPLATE_AUTO_RELOAD"] = True
app.config["SESSION_FILE_DIR"] = mkdtemp()
app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"
Session(app)

# Ensure responses aren't cached
@app.after_request
def after_request(response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Expires"] = 0
    response.headers["Pragma"] = "no-cache"
    return response

#Still use CS50 libary since i am not familiar with ORM
db = SQL("sqlite:///tracker.db")


def login_required(f):
    """
    Decorate routes to require login.
    http://flask.pocoo.org/docs/1.0/patterns/viewdecorators/
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("user_id") is None:
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated_function

def merge(list1, list2):

    merged_list = [(list1[i], list2[i]) for i in range(0, len(list1))]
    return merged_list

@app.route("/", methods = ["GET", "POST"])
@login_required
def index():
    user_id = session["user_id"]
    projects = db.execute("SELECT project,progress,pj_id FROM projects WHERE id = :user_id", user_id = user_id)
    employees = db.execute("SELECT employee FROM employees WHERE id=:user_id and status=:status;",user_id = user_id, status="idle")


    #manage data as needfor template
    i= 0
    task_list = []
    flag_list = []
    for i in range(len(projects)):
        task_rows = db.execute("SELECT task,flag FROM tasks WHERE pj_id = :pj_id", pj_id = projects[i]["pj_id"])
        emp_rows = db.execute("SELECT employee FROM employees WHERE pj_id = :pj_id ;", pj_id = projects[i]["pj_id"] )
        projects[i]['employee']= []
        if len(emp_rows) != 0:
            for emp in emp_rows:
                projects[i]['employee'].append(emp['employee'])

        if len(task_rows) != 0:
            for task in task_rows:
                task_list.append(task['task'])
                flag_list.append(task['flag'])

        projects[i]['tandf']=merge(task_list,flag_list)
        task_list.clear()
        flag_list.clear()
        i = i + 1

    for project in projects:
        for emp in project['employee']:
            print(emp)
        print()
        for t,f in project['tandf']:
            print(t,"\t",f)

        print(project['project'])
        print(project['progress'])
        print()


    if request.method == "POST":
        #adding task
        if request.form.get("clicked") == "add":
            pj_id = request.form.get("project_id")
            task = request.form.get("task")
            db.execute("INSERT INTO tasks (pj_id, task, flag) VALUES (:pj_id, :task, :flag);",
                        pj_id = pj_id, task = task, flag = "uncheck")
            row_tasks = db.execute("SELECT COUNT(task) AS count FROM tasks WHERE pj_id=:pj_id;",
                                pj_id = pj_id)
            row_checked = db.execute("SELECT COUNT(task) AS count FROM tasks WHERE flag = :flag AND pj_id = :pj_id;",
                                flag = "checked" , pj_id = pj_id)
            progress = int(100 * (int(row_checked[0]['count'])/int(row_tasks[0]['count'])))
            db.execute("UPDATE projects SET progress = :progress WHERE pj_id = :pj_id;", progress = progress, pj_id = pj_id)
            return redirect("/")

        #remove task
        if request.form.get("clicked") == "remove":
            pj_id = request.form.get("pj_id")
            task = request.form.get("t")
            db.execute("DELETE FROM tasks WHERE task LIKE :task AND pj_id = :pj_id;", task = task+"%", pj_id = pj_id)
            return redirect("/")

        if request.form.get("clicked") == "check":
            pj_id = request.form.get("pj_id")
            task = request.form.get("t")
            flag_stat = request.form.get("f")

            if flag_stat == "uncheck":
                db.execute("UPDATE tasks SET flag = :flag WHERE task LIKE :task and pj_id = :pj_id;",
                    flag = "checked", task = task+"%", pj_id = pj_id )
                row_tasks = db.execute("SELECT COUNT(task) AS count FROM tasks WHERE pj_id=:pj_id;",
                                pj_id = pj_id)
                row_checked = db.execute("SELECT COUNT(task) AS count FROM tasks WHERE flag = :flag AND pj_id = :pj_id;",
                                flag = "checked" , pj_id = pj_id)
                progress = int(100 * (int(row_checked[0]['count'])/int(row_tasks[0]['count'])))
                db.execute("UPDATE projects SET progress = :progress WHERE pj_id = :pj_id;", progress = progress, pj_id = pj_id)
                return redirect("/")

            elif flag_stat == "checked":
                db.execute("UPDATE tasks SET flag = :flag WHERE task LIKE :task and pj_id=:pj_id;",
                    flag = "uncheck", task = task+"%", pj_id = pj_id )
                row_tasks = db.execute("SELECT COUNT(task) AS count FROM tasks WHERE pj_id=:pj_id;",
                                pj_id = pj_id)
                row_checked = db.execute("SELECT COUNT(task) AS count FROM tasks WHERE flag = :flag AND pj_id = :pj_id;",
                                flag = "checked" , pj_id = pj_id)
                progress = int(100 * (int(row_checked[0]['count'])/int(row_tasks[0]['count'])))
                db.execute("UPDATE projects SET progress = :progress WHERE pj_id = :pj_id;", progress = progress, pj_id = pj_id)
                return redirect("/")

    else:
        return render_template("index.html",projects = projects ,employees =employees)


@app.route("/manage", methods = ["GET", "POST"])
@login_required
def manage():
    user_id = session["user_id"]
    employees = db.execute("SELECT employee FROM employees WHERE id=:user_id and status=:status;",user_id = user_id, status="idle")
    rows = db.execute("SELECT project,pj_id FROM projects WHERE id=:user_id;", user_id = user_id)

    if request.method == "POST":

        if request.form.get("clicked") == "create":
            project = request.form.get("project")
            team_tags = request.form.get("hidden-tags").split(",")
            if not project:
                return render_template("manage.html",c_error = "Project Name is required!")
            if not team_tags:
                return render_template("manage.html",c_error = "Need to add team members!")
            r = db.execute("SELECT pj_id FROM projects")
            db.execute("INSERT INTO projects (project,id,progress,pj_id) VALUES (:project, :user_id, :progress, :pj_id);",
                        project = project, user_id = user_id ,progress = 0, pj_id = len(r))
            pj_id = db.execute("SELECT pj_id FROM projects WHERE project = :project;", project = project)
            for employee in team_tags:
                print("Adding")
                db.execute("UPDATE employees SET pj_id = :pj_id , status=:status WHERE employee LIKE :employee;",
                            pj_id = pj_id[0]["pj_id"], employee = employee+"%", status = "active")
            return redirect("/manage")

        elif request.form.get("clicked") == "delete":
            project = request.form.get("pj")
            pj_id = request.form.get("pj_id")
            db.execute("DELETE FROM tasks WHERE pj_id = :pj_id;", pj_id = pj_id)
            db.execute("DELETE FROM projects WHERE pj_id= :pj_id;", pj_id = pj_id)
            db.execute("UPDATE employees SET pj_id = :pj_id,status = :status WHERE pj_id=:proj_id;",
                        pj_id = "NULL", status = "idle", proj_id = pj_id)
            return redirect("/manage")
        return redirect("/manage")
    else:
        return render_template("manage.html",employees = employees, rows = rows)

@app.route("/employees", methods = ["GET", "POST"])
@login_required
def employees():
    user_id = session["user_id"]
    rows = db.execute("SELECT employee,status FROM employees WHERE id= :user_id ORDER BY status;", user_id = user_id)
    if(request.method == "POST"):
        employee = request.form.get("employee")
        #add button
        if request.form["btn_clicked"] == "b1":
            if not employee:
                rows = db.execute("SELECT employee,status FROM employees WHERE id = :user_id ORDER BY status;", user_id = user_id)
                return render_template("employees.html", e_name = "Name is Empty", rows = rows)
            for row in rows:
                if(row["employee"] == employee):
                    rows = db.execute("SELECT employee,status FROM employees WHERE id= :user_id;", user_id = user_id)
                    return render_template("employees.html", e_name = "Employee is already added!", row = rows)
            db.execute("INSERT INTO employees (employee, id, status) VALUES (:employee ,:user_id, :status); ",
                        employee = employee , user_id = user_id, status = "idle")
            rows = db.execute("SELECT employee FROM employees WHERE id= :user_id ORDER BY status;", user_id = user_id)
            return redirect("/employees")

        #remove button
        elif request.form["btn_clicked"] == "b2":
            rows = db.execute("SELECT employee,status FROM employees WHERE id= :user_id ORDER BY status;", user_id = user_id)
            if not employee:
                return render_template("employees.html", e_name = "Name is Empty", rows = rows)
            for row in rows:
                if(row["employee"] == employee):
                    db.execute("DELETE FROM employees WHERE employee = :employee;", employee = employee)
                    return redirect("/employees")
            return render_template("employees.html", e_name = "Requested employee doesn't exist!" , rows = rows)

    else:
        return render_template("employees.html", rows= rows)




@app.route("/login", methods = ["GET", "POST"])
def login():
    if(request.method == "POST"):
        if not request.form.get("username"):
            return render_template("login.html", u_error = "Empty Username!")

        elif not request.form.get("password"):
            return render_template("login.html", p_error = "Empty Password")
        username = request.form.get("username")
        password = request.form.get("password")

        rows = db.execute("SELECT * from users where username = :username;", username = username)

        #check this user in table
        if len(rows) != 1 or not check_password_hash(rows[0]["hash"], password):
            return render_template("login.html", invalid = "Invalid username/password!")
        session["user_id"] = rows[0]['id']
        return redirect("/")
    else:
        return render_template("login.html")



@app.route("/logout")
def logout():

    session.clear()

    return redirect("/login")


@app.route("/chpwd", methods=["GET", "POST"])
@login_required
def chpwd():

    """Change Password"""
    if request.method == "POST":
        user_id = session["user_id"]
        if not request.form.get("old_password"):
            return render_template("chpwd.html", o_pwd = "Empty old password")
        elif not request.form.get("password"):
            return render_template("chpwd.html", n_pwd = "Empty new password")
        elif not request.form.get("password_cf"):
            return render_template("chpwd.html", rn_pwd = "Empty retype password")
        #check password and generate hash
        if(request.form.get("password") != (request.form.get("password_cf"))):
            # USE THE SAME PASSWORD
            return render_template("chpwd.html", rn_pwd = "Repeat new password")
        rows = db.execute("SELECT hash FROM users WHERE id = :user_id;",user_id=user_id)

        if len(rows) != 1 or not check_password_hash(rows[0]["hash"], request.form.get("old_password")):
            return render_template("chpwd.html", error = "Wrong old Password")

        hashpwd = generate_password_hash(request.form.get("password"), method="pbkdf2:sha256", salt_length=8)

        db.execute("UPDATE users SET hash = :hashpwd WHERE id = :user_id;", hashpwd = hashpwd, user_id = user_id)
        session.clear()
        return redirect("/login")
    else:
        return render_template("chpwd.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        if not request.form.get("username"):
            # will be replace with render template and use session data
            return render_template("register.html", u_error = "Empty Username!")
        elif not request.form.get("password"):
            return render_template("register.html", p_error = "Empty Password!")
        elif not request.form.get("password_cf"):
            return render_template("register.html", cf_error = "Repeat Password!")

        row = db.execute("SELECT * FROM users WHERE username = :username",
                         username = request.form.get("username"))

        #user already exists
        if len(row) == 1:
            return render_template("register.html", m_error = "User already exists" )

        #check password and generate hash
        if(request.form.get("password") != (request.form.get("password_cf"))):
            return  render_template("register.html", m_error = "Password mismatch")

        username = request.form.get("username")
        hashpwd = generate_password_hash(request.form.get("password"), method="pbkdf2:sha256", salt_length=8)

        okay = db.execute("INSERT INTO users (username, hash) VALUES (:username, :hash_val)", username = username, hash_val = hashpwd)
        return redirect("/login")
    else:
        return render_template("register.html")


@app.route("/export_report", methods=["GET"])
@login_required
def export_report():
    user_id = session["user_id"]
    projects = db.execute("SELECT project, progress, pj_id FROM projects WHERE id = :user_id", user_id=user_id)

    # Create a new Excel workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Report"

    # Set up the headers
    headers = ["Project Name", "Progress", "Task", "Task Status", "Employees"]
    ws.append(headers)

    # Style the headers
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Initialize the row counter
    row_counter = 2

    # Populate the Excel sheet with project data
    for project in projects:
        tasks = db.execute("SELECT task, flag FROM tasks WHERE pj_id = :pj_id", pj_id=project["pj_id"])
        employees = db.execute("SELECT employee FROM employees WHERE pj_id = :pj_id", pj_id=project["pj_id"])

        # Prepare task and employee lists
        emp_list = ', '.join([emp['employee'] for emp in employees]) if employees else "No employees assigned"

        if tasks:
            task_count = len(tasks)
            for task in tasks:
                task_name = task['task']
                task_status = "Completed" if task['flag'] == 'checked' else "Pending"
                ws.append([project['project'], f"{project['progress']}%", task_name, task_status, emp_list])
                row_counter += 1

            # Merge the project name and progress cells to span across all tasks
            ws.merge_cells(start_row=row_counter-task_count, start_column=1, end_row=row_counter-1, end_column=1)
            ws.merge_cells(start_row=row_counter-task_count, start_column=2, end_row=row_counter-1, end_column=2)

            # Align merged cells content to the center vertically
            for i in range(row_counter-task_count, row_counter):
                ws[f"A{i}"].alignment = Alignment(vertical="center")
                ws[f"B{i}"].alignment = Alignment(vertical="center")

        else:
            # If there are no tasks, avoid merging cells
            ws.append([project['project'], f"{project['progress']}%", "No tasks assigned", "", emp_list])
            row_counter += 1

    # Adjust column width for better readability
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Save the workbook to a BytesIO object
    output = BytesIO()
    wb.save(output)

    # Set the position of the stream to the beginning
    output.seek(0)

    # Send the file to the client
    return send_file(output, download_name="projects_report.xlsx", as_attachment=True)
